import pathlib

import scrapy
from os import remove
import sys, os
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.borders import Border, Side

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment



class QuotesSpider(scrapy.Spider):
    name = "quotes"
    page = 1

    def start_requests(self):
        # DIRECTORIO ACTUAL DONDE ESTAMOS
        ruta = str(pathlib.Path().resolve()) + "\demo.xlsx"
        wb = openpyxl.load_workbook(ruta)
        
        sheet_obj = wb.active
        m_row = sheet_obj.max_row

        #CREANDO EXCEL
        dest_filename = 'demo.xlsx'

        #ESTILOS PARA LA HOJA
        font = Font(bold=True)
        font_size = Font(size=16)

        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='06A30B',
                   fill_type='solid')

        greyFill = PatternFill(start_color='00C0C0C0',
                   end_color='06A30B',
                   fill_type='solid')
                   
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        ws1 = wb.active

        ws1.merge_cells('B3:C3')
        ws1['B3'] = "Ingrese un nombre y el link, el link que sea de Mediafire, Mega o Google Drive"
        ws1['B3'].font = font_size 

        #TITULO DE LAS COLUMAS
        ws1['A6'] = "Nombre"
        ws1['A6'].font = font
        ws1['A6'].border = thin_border
        ws1.cell(row = 6, column = 1).alignment = Alignment(horizontal='center')

        ws1['B6'] = "Link"
        ws1['B6'].font = font
        ws1['B6'].border = thin_border
        ws1.cell(row = 6, column = 2).alignment = Alignment(horizontal='center')

        ws1['C6'] = "Servidor"
        ws1['C6'].font = font
        ws1['C6'].border = thin_border
        ws1.cell(row = 6, column = 3).alignment = Alignment(horizontal='center')
        
        ws1['D6'] = "Estado"
        ws1['D6'].font = font
        ws1['D6'].border = thin_border
        ws1.cell(row = 6, column = 4).alignment = Alignment(horizontal='center')


        ws1.column_dimensions['E'].width = 150
        ws1.column_dimensions['F'].width = 14
     
        ws1.title = "Links"
        cont = 6   


        array_links = []

        for i in range(7, m_row + 1):
            cont+=1
        
            ws1.cell(row = i, column = 3).alignment = Alignment(horizontal='center') 
            ws1.cell(row = i, column = 4).alignment = Alignment(horizontal='center')     


            cell_obj = sheet_obj.cell(row = i, column = 2)
            url = cell_obj.value.strip()

            yield scrapy.Request(url=url, dont_filter=True, callback=self.parse, 
                meta={'cont': cont, 'ws1':ws1, 'greenFill': 
                greenFill, 'redFill': redFill, 'greyFill': greyFill, 'thin_border': thin_border,
                'dest_filename':dest_filename, 'wb':wb, 'array_links':array_links})
                #print("xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx")
               
        wb.save(filename = dest_filename)

  
    def parse(self, response):

        def NoDisponible():
            textoAc = "No Disponible"
            ws1['D'+ str(cont)].fill = redFill 
            ws1['D'+ str(cont)] = textoAc

        def Disponible():
            textoAc = "Disponible"
            ws1['D'+ str(cont)].fill = greenFill 
            ws1['D'+ str(cont)] = textoAc

        wb = response.meta.get('wb')
        ws1 = response.meta.get('ws1')
        dest_filename = response.meta.get('dest_filename')
        cont = response.meta.get('cont')
        greenFill = response.meta.get('greenFill')
        redFill = response.meta.get('redFill')
        greyFill = response.meta.get('greyFill')
        thin_border = response.meta.get('thin_border')
        array_links = response.meta.get('array_links')

        ws1['A'+ str(cont)].border = thin_border
        ws1['B'+ str(cont)].border = thin_border
        ws1['C'+ str(cont)].border = thin_border
        ws1['D'+ str(cont)].border = thin_border
       

        #Buscar en que servidor esta el archivo
        nombre = str(response.url)      

        #Buscar si el link esta disponible
        if nombre.find("mega") > -1:
            ws1['C'+ str(cont)] = "Mega"
            texto = response.xpath("//meta[@property='og:title']/@content").extract_first()
            if texto == "File on MEGA" or response.status == 404:
                NoDisponible()     
            else: 
                Disponible()
        elif nombre.find("mediafire") > -1:
            ws1['C'+ str(cont)] = "MediaFire"
            texto = response.xpath("//meta[@property='og:title']/@content").extract_first()
            if texto == "MediaFire" or response.status == 404:
                NoDisponible() 
            else: 
                Disponible()
        elif nombre.find("drive") > -1:
            ws1['C'+ str(cont)] = "Google Drive"
            texto = response.xpath("//meta[@property='og:url']/@content").extract_first()
            if texto == None or response.status == 404:
                NoDisponible() 
            else: 
                Disponible()

        #SABER SI EN LINK ESTA REPETIDO
        if nombre in array_links:
            ws1['D'+ str(cont)].fill = greyFill 
            ws1['D'+ str(cont)] = "Repetido"

        array_links.append(nombre)

        wb.save(filename = dest_filename)
        
 

        
 
